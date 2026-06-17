import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from jellyfish import jaro_winkler_similarity
import argparse

def parse_args():
    parser = argparse.ArgumentParser(description='Perform entity resolution on incident data.')
    parser.add_argument('--incident-data', required=True, help='Path to incident data CSV')
    parser.add_argument('--post-data', required=True, help='Path to POST data CSV')
    parser.add_argument('--output', required=True, help='Path to output Excel file')
    parser.add_argument('--training-output', required=True, help='Path to output training data CSV')
    return parser.parse_args()


def calculate_name_similarity(incident_name, post_name, component_type=''):
    """
    Calculate name similarity with enhanced matching for name variations.
    
    Args:
        incident_name: Name component from incident record
        post_name: Name component from POST record
        component_type: Type of name component ('first_name', 'last_name', 'middle_name', 'suffix')
        
    Returns:
        tuple: (similarity_score, should_include)
    """
    # Convert to strings and clean
    inc_name = str(incident_name).strip().lower() if pd.notna(incident_name) else ''
    post_name = str(post_name).strip().lower() if pd.notna(post_name) else ''
    
    # Both names are empty - don't include in similarity calculation
    if not inc_name and not post_name:
        return 0.0, False
        
    # One name is missing while the other isn't - don't include in similarity calc
    if bool(inc_name) != bool(post_name):
        return 0.0, False

    # Special handling for middle names/initials
    if component_type == 'middle_name':
        # If one is initial and matches start of other name
        if (len(inc_name) == 1 and post_name.startswith(inc_name)) or \
           (len(post_name) == 1 and inc_name.startswith(post_name)):
            return 1.0, True

    if component_type == 'suffix_name':
        # If one is initial and matches start of other name
        if (len(inc_name) == 1 and post_name.startswith(inc_name)) or \
           (len(post_name) == 1 and inc_name.startswith(post_name)):
            return 1.0, True
            
    # Do the Jaro-Winkler comparison 
    similarity = jaro_winkler_similarity(inc_name, post_name)
        
    return similarity, True

def calculate_weighted_similarity(similarities, weights):
    """
    Calculate weighted similarity score only including components that should be compared.
    
    Args:
        similarities: Dict of tuples (similarity_score, should_include)
        weights: Dict of component weights
        
    Returns:
        float: Weighted similarity score
    """
    weighted_sum = 0
    total_weight = 0
    
    for component, (similarity, should_include) in similarities.items():
        if should_include:  # Only include in calculation if should_include is True
            component_name = component.replace('_sim', '')
            weight = weights.get(component_name, 0)
            weighted_sum += similarity * weight
            total_weight += weight
    
    return weighted_sum / total_weight if total_weight > 0 else 0


def perform_entity_resolution_excel(incident_df, post_df, output_path='entity_matches.xlsx'):
    incident_df = incident_df.fillna("")
    post_df = post_df.fillna("")
    # List of target Fresno area agencies
    target_agencies = [
        "Fresno", "Selma", "Parlier", "Clovis", "Reedly", "Sanger", "Kerman",
        "Kingsburg", "Coalinga", "Firebaugh", "Calwa", "Friant", "Mendota",
        "Fowler", "Shaver Lake", "San Joaquin", "Orange Cove", "Del Ray",
        "Yokuts Valley", "Auberry", "Huron", "Caruthers", "Riverdale", "Laton",
        "Big Creek", "Tranquility", "Biola", "Raisin City", "Easton", "Three Rocks",
        "Cantou Creek", "Lanare", "Minkler", "Mayfair", "Malaga", "Bowles",
        "Monmouth", "West Park"
    ]
    
    # Define fixed weights
    weights = {
        'first_name': 0.45,
        'last_name': 0.45,
        'middle_name': 0.5,
        'suffix': 0.5
    }
    
    # Convert dates to datetime
    incident_df['incident_date'] = pd.to_datetime(incident_df['incident_date'])
    post_df['start_date'] = pd.to_datetime(post_df['start_date'])
    post_df['end_date'] = pd.to_datetime(post_df['end_date'])

    incident_df['incident_date'] = pd.to_datetime(incident_df['incident_date']).dt.strftime('%Y-%m-%d')
    post_df['start_date'] = pd.to_datetime(post_df['start_date']).dt.strftime('%Y-%m-%d')
    post_df['end_date'] = pd.to_datetime(post_df['end_date']).dt.strftime('%Y-%m-%d')
    
    # Filter POST records to only include rows where agency is in target list
    fresno_post = post_df[post_df['agency_name'].str.contains('|'.join(target_agencies), case=False, na=False)]

    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Track progress
        total_rows = len(incident_df)
        print(f"Processing {total_rows} incident records...")
        
        # For each incident record
        for idx, incident in incident_df.iterrows():
            if idx % 10 == 0:
                print(f"Processing row {idx}/{total_rows}")
                
            incident_date = incident['incident_date']
            incident_uid = incident['uid']
            
            # Get relevant POST records
            relevant_post = fresno_post[
                (fresno_post['start_date'] <= incident_date) & 
                ((fresno_post['end_date'] >= incident_date) | pd.isna(fresno_post['end_date']))
            ]
            
            # Track which components were present for incident record
            present_components = {
                'has_first': bool(str(incident['first_name']).strip()),
                'has_last': bool(str(incident['last_name']).strip()),
                'has_middle': bool(str(incident['middle_name']).strip()),
                'has_suffix': bool(str(incident['suffix']).strip())
            }
            
            # List to store all candidate matches
            candidate_matches = []
            
            for _, post in relevant_post.iterrows():
                # Calculate raw similarities without weights
                similarities = {
                    'first_name_sim': calculate_name_similarity(
                        incident['first_name'], 
                        post['first_name'],
                        'first_name'
                    ),
                    'last_name_sim': calculate_name_similarity(
                        incident['last_name'], 
                        post['last_name'],
                        'last_name'
                    ),
                    'middle_name_sim': calculate_name_similarity(
                        incident['middle_name'], 
                        post['middle_name'],
                        'middle_name'
                    ),
                    'suffix_sim': calculate_name_similarity(
                        incident['suffix'], 
                        post['suffix'],
                        'suffix'
                    )
                }
                
                # Store just the similarity scores for output
                similarity_scores = {k: v[0] for k, v in similarities.items()}
                
                # Calculate weighted average similarity using only relevant components
                weighted_sim = calculate_weighted_similarity(similarities, weights)
                
                # Calculate minimum required score based on available data
                min_required = 0.6
                if not (str(incident['first_name']).strip() and str(incident['last_name']).strip()):
                    min_required = 0.7
                
                # Create candidate match entry
                candidate_data = {
                    'incident_uid': incident_uid,
                    'post_person_nbr': post['person_nbr'],
                    'incident_name': f"{incident['first_name']} {incident['middle_name']} {incident['last_name']} {incident['suffix']}".strip(),
                    'incident_first_name': incident['first_name'],
                    'incident_middle_name': incident['middle_name'],
                    'incident_last_name': incident['last_name'],
                    'incident_suffix': incident['suffix'],
                    'post_name': f"{post['first_name']} {post['middle_name']} {post['last_name']} {post['suffix']}".strip(),
                    'post_first_name': post['first_name'],
                    'post_middle_name': post['middle_name'],
                    'post_last_name': post['last_name'],
                    'post_suffix': post['suffix'],
                    'incident_agency': incident['agency_name'],
                    'post_agency': post['agency_name'],
                    'incident_date': incident_date,
                    'post_start_date': post['start_date'],
                    'post_end_date': post['end_date'],
                    'similarity_score': weighted_sim,
                    'min_required_score': min_required,
                    **similarity_scores,
                    **present_components,
                    'components_present': sum(present_components.values())
                }
                
                candidate_matches.append(candidate_data)
            
            # Sort candidates by similarity score and take top 3
            candidate_matches.sort(key=lambda x: x['similarity_score'], reverse=True)
            top_candidates = candidate_matches[:3]
            
            # If we have fewer than 3 candidates, pad with empty matches
            while len(top_candidates) < 3:
                null_match = {
                    'incident_uid': incident_uid,
                    'post_person_nbr': None,
                    'incident_name': f"{incident['first_name']} {incident['middle_name']} {incident['last_name']} {incident['suffix']}".strip(),
                    'incident_first_name': incident['first_name'],
                    'incident_middle_name': incident['middle_name'],
                    'incident_last_name': incident['last_name'],
                    'incident_suffix': incident['suffix'],
                    'post_name': "No match found",
                    'post_first_name': "",
                    'post_middle_name': "",
                    'post_last_name': "",
                    'post_suffix': "",
                    'incident_agency': incident['agency_name'],
                    'post_agency': None,
                    'incident_date': incident_date,
                    'post_start_date': None,
                    'post_end_date': None,
                    'similarity_score': 0.0,
                    'min_required_score': min_required,
                    'first_name_sim': 0.0,
                    'last_name_sim': 0.0,
                    'middle_name_sim': 0.0,
                    'suffix_sim': 0.0,
                    **present_components,
                    'components_present': sum(present_components.values())
                }
                top_candidates.append(null_match)
            
            # Create DataFrame for this incident's matches
            matches_df = pd.DataFrame(top_candidates)
            
            # Select and reorder columns for display
            display_columns = [
                'incident_name', 'incident_first_name', 'incident_middle_name', 'incident_last_name', 'incident_suffix',
                'post_name', 'post_first_name', 'post_middle_name', 'post_last_name', 'post_suffix',
                'similarity_score',
                'first_name_sim', 'last_name_sim', 'middle_name_sim', 'suffix_sim',
                'incident_agency', 'post_agency',
                'incident_date', 'post_start_date', 'post_end_date',
                'incident_uid', 'post_person_nbr'
            ]
            
            matches_df = matches_df[display_columns]
            
            # Create sheet name (clean incident UID to be Excel-friendly)
            sheet_name = f"ID_{str(incident_uid)[:28]}"  # Excel has a 31 character limit for sheet names
            
            # Write to Excel
            matches_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Apply formatting
            worksheet = writer.sheets[sheet_name]
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
            
            # Color code similarity scores
            for row in range(2, 5):  # Rows 2-4 (1-based)
                score_cell = worksheet.cell(row=row, column=display_columns.index('similarity_score') + 1)
                score = float(score_cell.value or 0)
                
                if score >= 0.9:
                    score_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
                elif score >= 0.8:
                    score_cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Gold
                else:
                    score_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light red
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)
    
    print(f"Processing complete. Excel file created at: {output_path}")
    return output_path
def create_training_data(input_df, training_output_path, all_matches_path):
    selected_columns = [
        'incident_name',
        'incident_first_name',
        'incident_middle_name',
        'incident_last_name',
        'incident_suffix',
        'incident_date', 
        'post_name',
        'post_first_name',
        'post_middle_name',
        'post_last_name',
        'post_suffix',
        'post_start_date', 
        'post_end_date',
        'similarity_score',
        'incident_uid',
        'post_person_nbr',
        'incident_agency',
        'post_agency',
        'first_name_sim',
        'last_name_sim',
        'middle_name_sim',
        'suffix_sim'
    ]
    
    # Save complete dataset with all matches
    complete_df = input_df[selected_columns].copy()
    complete_df.to_csv(all_matches_path, index=False)
    print(f"Complete matches dataset saved to: {all_matches_path}")
    print(f"Complete dataset shape: {complete_df.shape}")
    
    # Create training dataset
    training_df = complete_df.copy()
    # training_df['label'] = (training_df['similarity_score'] == 1).astype(int)
    training_df.to_csv(training_output_path, index=False)
    
    print(f"Training data created and saved to: {training_output_path}")
    print(f"Training dataset shape: {training_df.shape}")
    # print(f"Number of positive cases (label=1): {training_df['label'].sum()}")
    # print(f"Number of negative cases (label=0): {(training_df['label'] == 0).sum()}")
    
    return training_df, complete_df

def main():
    args = parse_args()
    
    print(f"Reading incident data from {args.incident_data}")
    incident_df = pd.read_csv(args.incident_data)
    
    print(f"Reading POST data from {args.post_data}")
    post_df = pd.read_csv(args.post_data)
    
    print("Performing entity resolution...")
    matches_output = perform_entity_resolution_excel(
        incident_df=incident_df,
        post_df=post_df,
        output_path=args.output
    )
    
    matches_df = pd.read_excel(matches_output, sheet_name=None)
    all_matches = pd.concat(matches_df.values())
    
    # Define path for all matches CSV
    all_matches_path = args.output.rsplit('.', 1)[0] + '_all_matches.csv'
    
    # Create both training data and complete matches dataset
    training_df, complete_df = create_training_data(
        all_matches,
        args.training_output,
        all_matches_path
    )
    
    print(f"\nEntity resolution complete!")
    print(f"Excel results saved to: {matches_output}")
    print(f"All matches saved to: {all_matches_path}")
    print(f"Training data saved to: {args.training_output}")

if __name__ == "__main__":
    main()